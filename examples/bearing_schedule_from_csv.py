import pandas as pd
import openpyxl as xl
from datetime import datetime
# from gsapy import GSA


# # Define GSA model objects

# model_base = GSA('models\WBBn_base_v48_base.gwb')


# bearing_nodes = {
#     'reactions': [17027, 30098, 31307, 32316],
#     'displacements': [12110, 30097, 31306, 32315]
# }


# reactions = model_base.get_node_reactions(bearing_nodes['reactions'], 'A1')

# print(reactions)

























# Dictionary of input files
gsa_output_csv = {
    'static_cases': 'static_cases.csv', 
    'static_combinations': 'static_combinations.csv',
    'static_reactions': 'static_reactions.csv', 
    'static_displacements': 'static_displacements.csv',
    'traffic_cases': 'traffic_cases.csv',
    'traffic_displacements': 'traffic_displacements.csv',
    'traffic_reactions': 'traffic_reactions.csv'
    }

# ===========
# Functions
# ===========


def print_df_summary(df):
    '''
    Print summary dataframe head and tail 
    '''
    print(f'\nSummary of df:\n')
    print(f'{df.head()}\n...')
    print(f'{df.tail()}\n')


def export_to_excel(bearing_results):
    '''
    Export generated dataframes to excel

    Parameters:
    ------------  
    bearing_results: List of datframes with bearing result to export to excel
    '''
    timestamp = datetime.now()
    # Name of excel file to save bearing data
    output_xl = f'bearing_outputs_{timestamp.strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'

    # Write to output file
    with pd.ExcelWriter(output_xl, engine='openpyxl', mode = 'w') as writer:
        for df in bearings_results:
            displacements.to_excel(writer, sheet_name='displacements', index=False, )
            reactions.to_excel(writer, sheet_name='reactions', index=False)
            loads_and_combinations.to_excel(writer, sheet_name='loads', index=False)

    print('Successfully written outputs to excel')


def delete_ws_rows(ws):
    '''
    clear results from destination (dst) excel file using openpyxl.
    
    Continuously delete row 2 until there is only a single row left that
    contains column names
    '''
    while(ws.max_row > 1):
        # this removes row 2
        ws.delete_ws_rows(2)
    # return to main function
    print(f'\nDeleted old rows from {ws}')
    
def copy_ws(ws_src, ws_dst):
    '''
    Copy values from source worksheet (ws_src) to destination worsksheet (ws_dst).
    Does not copy header row, idx=0
    '''
    for row in ws_src:
        if row == 0:
            pass
        else:
            for cell in row:
                ws_dst[cell.coordinate].value = cell.value
    wb_dst.save(dst)


# import data using gsapy





























# =====================
# Generate DataFrames
# =====================

# Load Cases and combinations 
# ---------------------------
# Load and clean dataframes
static_cases = pd.read_csv(gsa_output_csv['static_cases'], skiprows=46, header=0, nrows=45, usecols=[1, 2])
static_cases.reset_index(drop=True, inplace=True)



traffic_cases = pd.read_csv(gsa_output_csv['traffic_cases'], skiprows=45, header=0, nrows=73, usecols=[1, 2])
traffic_cases.reset_index(drop=True, inplace=True)
traffic_cases['Case'] = 'Traffic_' + traffic_cases['Case']

# combine load case and combination dataframes
load_cases = pd.concat([static_cases, traffic_cases])
load_cases.reset_index(inplace=True, drop=True)

# ## Import combinations from gsa .csv files
combinations = pd.read_csv(gsa_output_csv['static_combinations'], skiprows=18, header=0, usecols=[0, 1])
combinations.dropna(axis=0, subset=['Name'], inplace=True)

# ## Join load cases and combinations into one dataframe
loads_and_combinations = pd.concat([load_cases, combinations])

# Reactions
# ---------
# Load and clean dataframes
static_reactions = pd.read_csv(gsa_output_csv['static_reactions'], skiprows=19)
static_reactions.dropna(subset=['Case'], inplace=True)
static_reactions['Node'] = pd.to_numeric(static_reactions['Node'])

traffic_reactions = pd.read_csv(gsa_output_csv['traffic_reactions'], skiprows=18)
traffic_reactions.dropna(subset=['Case'], inplace=True)
traffic_reactions['Node'] = pd.to_numeric(traffic_reactions['Node'])
traffic_reactions['Case'] = 'Traffic_' + traffic_reactions['Case']

reactions = pd.concat([static_reactions, traffic_reactions])

# Displacements
# ----------------
# load and clean dataframes
static_displacements = pd.read_csv(gsa_output_csv['static_displacements'], skiprows=19)
static_displacements.dropna(subset=['Case'], inplace=True)
static_displacements['Node'] = pd.to_numeric(static_displacements['Node'])

traffic_displacements = pd.read_csv(gsa_output_csv['traffic_displacements'], skiprows=18)
traffic_displacements.dropna(subset=['Case'], inplace=True)
traffic_displacements['Node'] = pd.to_numeric(traffic_displacements['Node'])
traffic_displacements['Case'] = 'Traffic_' + traffic_displacements['Case']

displacements = pd.concat([static_displacements, traffic_displacements])

# List with all dataframes
bearings_results = [loads_and_combinations, displacements, reactions]


# =================================
# Copy output to bearings schedule
# =================================
# Copy worksheets from outputs excel to bearing schedule workbook

# define source and destination
# results output file
# src = output_xl

# # bearing_schedule file
# dst = 'bearing_specification_v3.xlsx'

# # load source workbooks and worksheets
# wb_src = xl.load_workbook(filename=src)
# ws_src_displacements = wb_src['displacements']
# ws_src_reactions = wb_src['reactions']
# ws_src_loads = wb_src['loads']

# # load destination workbooks and worksheets
# wb_dst = xl.load_workbook(filename=dst)
# ws_dst_displacements = wb_dst['displacements']
# ws_dst_reactions = wb_dst['reactions']
# ws_dst_loads = wb_dst['loads']




# ==========================
# Generate one df with all data
# ==========================


# keys = load_types['Cases'].tolist()
# cases = []

# for x in range(len(load_types)):
#     start = int(load_types.loc[x, ['Start']])
#     end = int(load_types.loc[x, ['End']])
#     load = load_types.at[x, 'type']

#     cases.append([f'{load}{start}' for start in range(start, end+1)])

# cases_dict = dict(zip(keys, cases))

# print(cases_dict)


if __name__ == '__main__':

    print_df_summary(displacements)
    print_df_summary(reactions)
    print_df_summary(loads_and_combinations)

    export_to_excel(bearings_results)

    # delete_ws_rows(ws_dst_displacements)
    # delete_ws_rows(ws_dst_reactions)
    # delete_ws_rows(ws_dst_loads)

    # copy_ws(ws_src_displacements, ws_dst_displacements)
    # copy_ws(ws_src_reactions, ws_dst_reactions)
    # copy_ws(ws_src_loads, ws_dst_loads)

    # print('\nOutputs successfully copied to bearing schedule')
