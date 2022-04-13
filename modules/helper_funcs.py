import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
import openpyxl


"""
Collection of functions to use in main code
"""


def move_column_df(df, position, column_name):
    '''
    shift column_name to 'n' position in df
    '''
    column_to_move = df.pop(column_name)
    # insert column using insert(position,column_name,
    # first_column) function
    df.insert(position, column_name, column_to_move)
    return df


def write_to_excel(df_dict, filename, path=None):
    '''
    write dictionary of dataframes to excel.

    parameters:
    df_dict (dict): dictionary of dataframes
    filename (str): name of output file
    '''
    wb = Workbook()
    timestamp = datetime.now()
    # Name of excel file to save bearing data
    filename = f'{filename}_{timestamp.strftime("%Y-%m-%d_%H-%M")}.xlsx'
    wb.save(filename)
    for key, df in df_dict.items():
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=key)
    wb = openpyxl.load_workbook(filename)
    sheet_to_del = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(sheet_to_del)
    wb.
    wb.save(filename)
    print(f'\nSuccessfully written {filename} to excel')


def delete_ws_rows(ws):
    '''
    clear results from destination (dst) excel file using openpyxl.
    
    Continuously delete row 2 until there is only a single row left that
    contains column names

    Parameters:
    ws: openpyxl worksheet
    '''
    while(ws.max_row > 1):
        # this removes row 2
        ws.delete_ws_rows(2)
    # return to main function
    print(f'\nDeleted old rows from {ws}')


def copy_ws(ws_src, ws_dst):
    '''
    Copy values from source worksheet (ws_src) to destination worsksheet (ws_dst).
    Does not copy header row, idx=0
    '''
    for row in ws_src:
        if row == 0:
            pass
        else:
            for cell in row:
                ws_dst[cell.coordinate].value = cell.value
    wb_dst.save(dst)
